from openpyxl import load_workbook 
from datetime import datetime


def format_date(date_input):
    try:
        return datetime.strptime(date_input, "%y.%m.%d").strftime("%Y-%m-%d")
    except ValueError:
        raise ValueError(f"날짜 형식을 인식할 수 없습니다: {date_input}")


def update_excel(file_path, waybill_number, shipment_weight, pickup_date, ready_time, blood_collection_date):
    book = load_workbook(file_path)
    sheet = book.active

    shipment_weight_num = float(shipment_weight)
    temperature_status = 'F' if shipment_weight_num >= 4.1 else 'A' if shipment_weight_num == 1.0 else 'Unknown'

    sheet['B5'] = datetime.now().date()
    sheet['D5'] = waybill_number
    sheet['E5'] = temperature_status
    sheet['K5'] = format_date(pickup_date)
    sheet['L5'] = '12:00' if ready_time == '오전' else '16:00' if ready_time == '오후' else ready_time
    sheet['M5'] = format_date(blood_collection_date)
    
    book.save(file_path)
    print("엑셀 파일이 성공적으로 업데이트 되었습니다.")
    book.close()

def read_pickup_date(file_path='C:\Python\bill\코반스 픽업요청서 양식.xlsx'):
    book = load_workbook(file_path)
    sheet = book.active

    date = str(sheet['K5'].value)
    return date

def read_excel(file_path):
    # House Air Bill #, Expected Date of delivery 불러오기
    book = load_workbook(file_path)
    sheet = book.active

    billNum = sheet['D5'].value
    date = list(str(sheet['K5'].value).split('-'))
    print('date:', date)
    yy = date[0]
    mm = date[1]
    dd = date[2]

    mm_dict = {
        '01': 'JAN',
        '02': 'FEB',
        '03': 'MAR',
        '04': 'APR',
        '05': 'MAY',
        '06': 'JUN',
        '07': 'JUL',
        '08': 'AUG',
        '09': 'SEP',
        '10': 'OCT',
        '11': 'NOV',
        '12': 'DEC',
    }

    return billNum, '{}{}{}'.format(dd, mm_dict[mm], yy)



 

# if __name__ == "__main__":
#     main()

